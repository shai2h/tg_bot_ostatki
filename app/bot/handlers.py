from __future__ import annotations

import asyncio
import logging
import os
import re
from decimal import Decimal, InvalidOperation
from zoneinfo import ZoneInfo

import pandas as pd
from aiogram import F
from aiogram.filters import Command
from aiogram.types import (
    CallbackQuery,
    FSInputFile,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy import select

from app.bot.user_guards import (
    inbound_callback_key,
    inbound_message_key,
    is_message_from_bot,
    user_guards,
)
from app.bot.utils import format_stock_quantity
from app.config import settings
from app.db.database import async_session_maker
from app.services.catalog_client import search_products as b2b_search_products
from app.services.catalog_exceptions import (
    CatalogClientError,
    CatalogConfigurationError,
    CatalogRateLimitError,
    CatalogResponseError,
    CatalogUnauthorizedError,
    CatalogUnavailableError,
    CatalogValidationError,
)
from app.services.catalog_models import CatalogProduct, CatalogSearchResponse
from app.services.search import (
    find_products_by_text as find_products_by_query,
    fuzzy_find_products,
    get_user_query_history,
    log_user_query,
)
from app.warehouse_stock.models import OstatkiMeta, WarehouseStocks

logger = logging.getLogger(__name__)


_HTML_TAG_RE = re.compile(r"</?(?:b|strong|i|em|u|s|code|pre|a)(?:\s+[^>]*)?>", re.IGNORECASE)
PRICE_LIST_URL = "https://rosholod.org/price-lists/OstatkiPoStolbcam.xls"
PRICE_LIST_TEXT = (
    "\u0410\u043a\u0442\u0443\u0430\u043b\u044c\u043d\u044b\u0439 "
    "\u043f\u0440\u0430\u0439\u0441 \u043c\u043e\u0436\u043d\u043e "
    "\u0441\u043a\u0430\u0447\u0430\u0442\u044c \u043f\u043e "
    f"\u0441\u0441\u044b\u043b\u043a\u0435:\n{PRICE_LIST_URL}"
)
PRICE_LIST_BUTTON_TEXT = "\u0421\u043a\u0430\u0447\u0430\u0442\u044c \u043f\u0440\u0430\u0439\u0441"
DEALER_PLATFORM_URL = "https://rosholod.org"
CARD_SEPARATOR = "━━━━━━━━━━━━━━━━━━"

WELCOME_TEXT = (
    "👋 Добро пожаловать!\n\n"
    "Здесь вы можете быстро узнать наличие товаров на складах Росхолода.\n\n"
    "🔎 Отправьте название товара, артикул или код товара — "
    "и я покажу актуальную информацию о наличии.\n\n"
    "🌐 Работайте быстрее через дилерскую платформу Росхолод.\n\n"
    "На платформе вы можете:\n"
    "• оформлять заказы онлайн;\n"
    "• проверять актуальные цены и остатки;\n"
    "• просматривать документы и историю заказов.\n\n"
    "👉 Для регистрации обратитесь к вашему персональному менеджеру.\n\n"
    f"🔗 {DEALER_PLATFORM_URL}"
)

DEALER_PROMO_TEXT = (
    f"{CARD_SEPARATOR}\n\n"
    "🌐 Пользуетесь ботом регулярно?\n\n"
    "Попробуйте дилерскую платформу Росхолод.\n\n"
    "✔️ Оформление заказов онлайн\n"
    "✔️ Актуальные остатки и цены\n"
    "✔️ Документы и история заказов\n\n"
    "👉 Для подключения обратитесь к вашему персональному менеджеру.\n\n"
    f"🔗 {DEALER_PLATFORM_URL}"
)

_AVAILABILITY_EMOJI = (
    ("немного", "🟠"),
    ("в наличии", "🟡"),
    ("много", "🟢"),
    ("нет", "🔴"),
)
STATUS_EMOJI = {
    "many": "🟢",
    "available": "🟡",
    "few": "🟠",
}
STATUS_LABEL_FALLBACK = {
    "many": "Много",
    "available": "В наличии",
    "few": "Немного",
}
POSITIVE_AVAILABILITY_STATUSES = set(STATUS_EMOJI)
NO_STOCK_MESSAGE = "На складах пока нет доступного остатка."


def _price_list_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text=PRICE_LIST_BUTTON_TEXT,
                    url=PRICE_LIST_URL,
                )
            ]
        ]
    )


def _url_button_keyboard(text: str, url: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text=text, url=url)]]
    )


def _normalize_query(query: str) -> str:
    return " ".join(query.strip().split())


def _availability_emoji(label: str | None, status: str | None = None) -> str:
    status_emoji = STATUS_EMOJI.get((status or "").strip())
    if status_emoji:
        return status_emoji

    needle = f"{label or ''} {status or ''}".casefold()
    for token, emoji in _AVAILABILITY_EMOJI:
        if token in needle:
            return emoji
    return "⚪"


def _format_availability_label(label: str | None, status: str | None = None) -> str:
    text = (label or "").strip() or "Наличие уточняется"
    return f"{_availability_emoji(text, status)} {text}"


def _format_product_availability_lines(product: CatalogProduct) -> list[str]:
    if product.warehouses:
        lines = ["Наличие:"]
        for warehouse in product.warehouses:
            status = (warehouse.status or "").strip()
            if status not in POSITIVE_AVAILABILITY_STATUSES:
                continue
            display_city = ((warehouse.city or warehouse.name or "")).strip()
            if not display_city:
                continue
            label = (warehouse.label or "").strip() or STATUS_LABEL_FALLBACK[status]
            lines.append(f"{STATUS_EMOJI[status]} {display_city} — {label}")
        if len(lines) > 1:
            return lines

    return [f"⚪ {NO_STOCK_MESSAGE}"]


def _format_retail_price_value(product: CatalogProduct) -> str | None:
    display = (product.retail_price_display or "").strip()
    if display:
        return display

    if product.retail_price is None:
        return None

    try:
        value = Decimal(product.retail_price)
    except (InvalidOperation, TypeError, ValueError):
        return None

    sign = "-" if value < 0 else ""
    absolute = abs(value)
    quantized = f"{absolute:.2f}"
    integer_part, _, fraction = quantized.partition(".")
    grouped = f"{int(integer_part):,}".replace(",", " ")
    if fraction.rstrip("0"):
        formatted = f"{grouped},{fraction.rstrip('0')}"
    else:
        formatted = grouped
    return f"{sign}{formatted} ₽"


def _append_optional_field(lines: list[str], label: str, value: str | None) -> None:
    if value is None:
        return
    text = str(value).strip()
    if not text:
        return
    lines.append(f"▶️ {label}: {text}")


def _format_exact_product_card(product: CatalogProduct) -> str:
    lines = [f"🧺 {product.title}", "", CARD_SEPARATOR, ""]
    _append_optional_field(lines, "Код", product.code)
    _append_optional_field(lines, "Артикул", product.article)
    _append_optional_field(lines, "Бренд", product.brand)
    _append_optional_field(lines, "Категория", product.category)
    price = _format_retail_price_value(product)
    if price:
        lines.append(f"▶️ Цена: {price}")

    lines.append("")
    lines.extend(_format_product_availability_lines(product))

    return "\n".join(lines).strip()


def _format_compact_product_card(index: int, product: CatalogProduct) -> str:
    lines = [f"🧺 {index}. {product.title}", "", CARD_SEPARATOR, ""]
    _append_optional_field(lines, "Код", product.code)
    _append_optional_field(lines, "Артикул", product.article)
    _append_optional_field(lines, "Бренд", product.brand)
    _append_optional_field(lines, "Категория", product.category)
    price = _format_retail_price_value(product)
    if price:
        lines.append(f"▶️ Цена: {price}")

    lines.append("")
    lines.extend(_format_product_availability_lines(product))

    return "\n".join(lines).strip()


def _record_b2b_fallback() -> None:
    try:
        from app.observability.prometheus_metrics import b2b_catalog_search_fallback_total

        b2b_catalog_search_fallback_total.inc()
    except Exception:  # pragma: no cover
        logger.debug("B2B fallback metric skipped", exc_info=True)


def _is_max_event(event: Message | CallbackQuery) -> bool:
    return getattr(event, "platform", "") == "max"


def _plain_text_for_max(text: str) -> str:
    text = text.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
    return _HTML_TAG_RE.sub("", text)


async def _answer(message: Message, text: str, **kwargs) -> None:
    if _is_max_event(message):
        text = _plain_text_for_max(text)
        kwargs.pop("parse_mode", None)
    await message.answer(text, **kwargs)


async def _answer_document(message: Message, document: FSInputFile, **kwargs) -> None:
    caption = kwargs.get("caption")
    if caption and _is_max_event(message):
        kwargs["caption"] = _plain_text_for_max(caption)
    await message.answer_document(document, **kwargs)


def _main_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="Инструкция"),
                KeyboardButton(text=PRICE_LIST_BUTTON_TEXT),
                KeyboardButton(text="История запросов"),
            ]
        ],
        resize_keyboard=True,
    )


def _main_menu_inline_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Инструкция", callback_data="menu:instruction")],
            [InlineKeyboardButton(text=PRICE_LIST_BUTTON_TEXT, url=PRICE_LIST_URL)],
            [InlineKeyboardButton(text="История запросов", callback_data="menu:history")],
            [InlineKeyboardButton(text="Открыть платформу", url=DEALER_PLATFORM_URL)],
        ]
    )


async def _get_last_updated_label() -> str:
    async with async_session_maker() as session:
        result = await session.execute(select(OstatkiMeta.last_updated).limit(1))
        updated_at = result.scalar()

    if not updated_at:
        return "неизвестно"

    return updated_at.astimezone(ZoneInfo("Europe/Moscow")).strftime("%d.%m.%Y %H:%M:%S")


async def _maybe_send_dealer_promo(message: Message, user_id: int, *, success: bool) -> None:
    if not user_guards.end_search(user_id, success=success):
        return
    await _answer(message, DEALER_PROMO_TEXT)


async def _run_search_with_thinking(message: Message, awaitable):
    task = asyncio.create_task(awaitable)
    try:
        try:
            return await asyncio.wait_for(asyncio.shield(task), timeout=1.0)
        except asyncio.TimeoutError:
            try:
                await _answer(message, "🔎 Ищу товар...")
            except Exception:
                logger.debug("Failed to send search thinking indicator", exc_info=True)
            return await task
    except BaseException:
        if not task.done():
            task.cancel()
            try:
                await task
            except (asyncio.CancelledError, Exception):
                pass
        raise


def _accept_inbound_message(message: Message) -> bool:
    """Drop bot echoes and duplicated message_created updates before any handler logic."""
    if is_message_from_bot(message):
        logger.info(
            "Ignoring bot-authored inbound message message_key=%s",
            inbound_message_key(message),
        )
        return False
    message_key = inbound_message_key(message)
    if user_guards.should_skip_duplicate_message(message_key):
        logger.info("Ignoring duplicated inbound message message_key=%s", message_key)
        return False
    return True


async def _accept_inbound_callback(callback: CallbackQuery) -> bool:
    callback_key = inbound_callback_key(callback)
    if user_guards.should_skip_duplicate_callback(callback_key):
        logger.info("Ignoring duplicated callback callback_key=%s", callback_key)
        try:
            await callback.answer()
        except Exception:
            logger.debug("callback.answer on duplicate failed", exc_info=True)
        return False
    return True


def register_handlers(router) -> None:
    @router.message(Command("start"))
    async def handle_start(message: Message) -> None:
        if not _accept_inbound_message(message):
            return
        user_id = message.from_user.id
        # Root cause of duplicates: MAX may deliver /start more than once, and the
        # catch-all F.text handler previously also accepted command texts.
        if user_guards.should_skip_duplicate_start(user_id):
            logger.info("Skipping duplicate /start user_id=%s", user_id)
            return

        await _answer(
            message,
            WELCOME_TEXT,
            reply_markup=_main_menu_inline_keyboard() if _is_max_event(message) else _main_menu_keyboard(),
        )

    @router.message(F.text.casefold() == "инструкция")
    async def handle_instruction(message: Message) -> None:
        if not _accept_inbound_message(message):
            return
        await _send_instruction(message)

    @router.message(F.text.casefold() == "полный отчет xlsx")
    @router.message(F.text.casefold() == "\u0441\u043a\u0430\u0447\u0430\u0442\u044c \u043f\u0440\u0430\u0439\u0441")
    async def handle_full_report(message: Message) -> None:
        if not _accept_inbound_message(message):
            return
        await _send_price_list(message)

    @router.message(F.text.casefold() == "история запросов")
    async def handle_history_request(message: Message) -> None:
        if not _accept_inbound_message(message):
            return
        user_id = message.from_user.id
        history = await get_user_query_history(user_id)
        if not history:
            await _answer(message, "История пуста.")
            return

        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text=query, callback_data=f"history:{query}")]
                for query in history[:5]
            ]
        )
        await _answer(message, "Выберите запрос из истории:", reply_markup=keyboard)

    @router.callback_query(F.data.startswith("history:"))
    async def handle_history_callback(callback: CallbackQuery) -> None:
        if not await _accept_inbound_callback(callback):
            return
        query = callback.data.replace("history:", "", 1)
        await callback.answer()
        await _send_search_results(callback.message, callback.from_user.id, query)

    @router.callback_query(F.data == "menu:instruction")
    async def handle_menu_instruction(callback: CallbackQuery) -> None:
        if not await _accept_inbound_callback(callback):
            return
        await callback.answer()
        await _send_instruction(callback.message, reply_markup=_main_menu_inline_keyboard())

    @router.callback_query(F.data == "menu:history")
    async def handle_menu_history(callback: CallbackQuery) -> None:
        if not await _accept_inbound_callback(callback):
            return
        await callback.answer()
        history = await get_user_query_history(callback.from_user.id)
        if not history:
            await _answer(callback.message, "История пуста.", reply_markup=_main_menu_inline_keyboard())
            return

        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text=query, callback_data=f"history:{query}")]
                for query in history[:5]
            ]
        )
        await _answer(callback.message, "Выберите запрос из истории:", reply_markup=keyboard)

    @router.callback_query(F.data == "menu:report")
    async def handle_menu_report(callback: CallbackQuery) -> None:
        if not await _accept_inbound_callback(callback):
            return
        await callback.answer()
        await _send_price_list(callback.message)

    @router.message(F.text)
    async def handle_user_query(message: Message) -> None:
        if not _accept_inbound_message(message):
            return
        text = (message.text or "").strip()
        # Prevent command texts (/start and others) from entering search flow.
        if text.startswith("/"):
            return
        await _send_search_results(message, message.from_user.id, text)


async def _send_price_list(message: Message) -> None:
    await _answer(message, PRICE_LIST_TEXT, reply_markup=_price_list_keyboard())
    return

    # Legacy XLSX generation is intentionally kept below for possible reuse.
    # It is disabled because the current flow sends a static price-list link.
    async with async_session_maker() as session:
        result = await session.execute(WarehouseStocks.__table__.select())
        rows = result.fetchall()

    if not rows:
        await _answer(message, "Нет данных для отчета.")
        return

    all_cities = sorted({row.sklad for row in rows})
    grouped: dict[tuple, dict[str, str]] = {}

    for row in rows:
        key = (row.vid, row.name, row.price, row.brend, row.kod, row.articul)
        grouped.setdefault(key, {city: "" for city in all_cities})
        grouped[key][row.sklad] = format_stock_quantity(row.ostatok)

    data = []
    for key, city_stocks in grouped.items():
        row_data = {
            "Вид номенклатуры": key[0],
            "Наименование": key[1],
            "Розничная цена (₽)": key[2],
            "Бренд": key[3],
            "Код": key[4],
            "Артикул": key[5],
        }
        row_data.update(city_stocks)
        data.append(row_data)

    latest_date = await _get_last_updated_label()
    safe_date = latest_date.replace(":", "-").replace(" ", "_")
    file_path = f"report_{safe_date}.xlsx"

    try:
        pd.DataFrame(data).to_excel(file_path, index=False)

        wb = load_workbook(file_path)
        ws = wb.active
        ws["A1"] = f"Актуальность остатков: {latest_date}"
        fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

        for col in ws.iter_cols(min_row=1, max_row=1):
            header = col[0].value
            col_letter = col[0].column_letter
            if header in ["Вид номенклатуры", "Наименование", "Бренд"]:
                ws.column_dimensions[col_letter].width = 40
            elif header in ["Код", "Артикул"]:
                ws.column_dimensions[col_letter].width = 20
            elif str(header).startswith("Розничная"):
                ws.column_dimensions[col_letter].width = 18
            else:
                ws.column_dimensions[col_letter].width = 15
                col[0].fill = fill

        wb.save(file_path)
        await _answer_document(
            message,
            FSInputFile(file_path),
            caption="Полный отчет по складам",
        )
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)


async def _send_search_results(message: Message, user_id: int, query: str) -> None:
    block_reason = user_guards.begin_search(user_id)
    if block_reason:
        await _answer(message, block_reason)
        return

    success = False
    try:
        await log_user_query(user_id, query)
        normalized_query = _normalize_query(query)

        if settings.B2B_CATALOG_SEARCH_ENABLED:
            try:
                response = await _run_search_with_thinking(
                    message,
                    b2b_search_products(normalized_query, limit=5),
                )
                await _send_b2b_search_results(message, normalized_query, response)
                success = True
                return
            except CatalogValidationError:
                await _answer(
                    message,
                    "Введите не менее 3 символов: название, код или артикул товара.",
                )
                return
            except (
                CatalogConfigurationError,
                CatalogUnauthorizedError,
                CatalogRateLimitError,
                CatalogUnavailableError,
                CatalogResponseError,
                CatalogClientError,
            ) as exc:
                _record_b2b_fallback()
                logger.warning(
                    "B2B search fallback to local search_source=local_fallback "
                    "query_length=%s error_type=%s",
                    len(normalized_query),
                    type(exc).__name__,
                )
                try:
                    await _run_search_with_thinking(
                        message,
                        _send_local_search_results(
                            message,
                            query,
                            search_source="local_fallback",
                        ),
                    )
                    success = True
                except Exception:
                    logger.exception(
                        "Local search fallback failed search_source=local_fallback "
                        "query_length=%s",
                        len(normalized_query),
                    )
                    await _answer(
                        message,
                        "Сервис поиска временно недоступен. Попробуйте ещё раз немного позже.",
                    )
                return

        await _run_search_with_thinking(
            message,
            _send_local_search_results(message, query, search_source="local"),
        )
        success = True
    finally:
        await _maybe_send_dealer_promo(message, user_id, success=success)


async def _send_b2b_search_results(
    message: Message,
    query: str,
    response: CatalogSearchResponse,
) -> None:
    products = response.products[:5]

    if not products:
        keyboard = None
        if response.catalog_search_url:
            keyboard = _url_button_keyboard(
                "Открыть каталог",
                response.catalog_search_url,
            )
        await _answer(
            message,
            (
                f"🤷 По запросу «{query}» ничего не найдено. \n\n"
                "Проверьте код, артикул или попробуйте часть названия.\n\n"
                "🌐 Посмотреть полный каталог товаров"
            ),
            reply_markup=keyboard,
        )
        return

    if response.exact_match:
        product = products[0]
        # MAX/obabot strips HTML/Markdown links, so keep a stable URL-button CTA.
        keyboard = None
        if product.product_url:
            keyboard = _url_button_keyboard("Открыть карточку товара", product.product_url)
        await _answer(message, _format_exact_product_card(product), reply_markup=keyboard)
        return

    if len(products) == 1:
        intro = "Найдено товаров: 1."
    elif len(products) < 5:
        intro = (
            f"🕵️ По запросу «{query}» найдено несколько товаров.\n"
            f"Найдено товаров: {len(products)}."
        )
    else:
        intro = (
            f" 🕵️По запросу «{query}» найдено несколько товаров.\n"
            "Показываю первые 5."
        )
    await _answer(message, intro)

    for index, product in enumerate(products, start=1):
        keyboard = None
        if product.product_url:
            keyboard = _url_button_keyboard("Открыть карточку товара", product.product_url)
        await _answer(
            message,
            _format_compact_product_card(index, product),
            reply_markup=keyboard,
        )

    if response.catalog_search_url:
        await _answer(
            message,
            "🌐 Посмотреть полный каталог товаров",
            reply_markup=_url_button_keyboard(
                "Открыть каталог",
                response.catalog_search_url,
            ),
        )


async def _send_local_search_results(
    message: Message,
    query: str,
    *,
    search_source: str,
) -> None:
    logger.info(
        "Local catalog search started search_source=%s query_length=%s",
        search_source,
        len(query),
    )
    items = await find_products_by_query(query)

    if items:
        latest_date = await _get_last_updated_label()
        logger.info(
            "Local catalog search ok search_source=%s result_count=%s",
            search_source,
            len(items),
        )

        if len(items) > 20:
            file_name = f"Результаты_{query.replace(' ', '_')}.txt"
            try:
                lines = []
                for product in items.values():
                    text = (
                        f"{product['name']}\n"
                        f"  Вид: {product['vid']}\n"
                        f"  Бренд: {product['brend']}\n"
                        f"  Артикул: {product['articul']}\n"
                        f"  Код: {product['kod']}\n"
                        f"  Цена: {product['price']} ₽\n"
                        "  Наличие:\n"
                    )
                    for stock in product["stocks"]:
                        text += f"    - {stock['sklad']}: {format_stock_quantity(stock['ostatok'])}\n"
                    text += "\n"
                    lines.append(text)

                with open(file_name, "w", encoding="utf-8") as file:
                    file.write("".join(lines))

                await _answer_document(
                    message,
                    FSInputFile(file_name),
                    caption=f"Найдено {len(items)} товаров по запросу: {query}",
                )
            finally:
                if os.path.exists(file_name):
                    os.remove(file_name)
            return

        for product in items.values():
            sklad_lines = "\n".join(
                f"• {stock['sklad']}: <b>{format_stock_quantity(stock['ostatok'])}</b>"
                for stock in product["stocks"]
            )
            text = (
                f"📦 <b>{product['name']} | {product['kod']}</b>\n"
                f"🏷️ <b>Бренд:</b> {product['brend']}\n"
                f"📌 <b>Вид:</b> {product['vid']}\n"
                f"🔖 <b>Артикул:</b> {product['articul'] or '-'}\n"
                f"💰 <b>Цена:</b> {product['price']} ₽\n\n"
                f"🚚 <b>Остатки по складам:</b>\n{sklad_lines}\n\n"
                f"🕒 <i>Актуально на: {latest_date}</i> по МСК"
            )
            await _answer(message, text)
        return

    fuzzy_results = await fuzzy_find_products(query)
    filtered = [result for result in fuzzy_results if result["score"] > 70]

    if not filtered:
        logger.info(
            "Local catalog search empty search_source=%s query_length=%s",
            search_source,
            len(query),
        )
        await _answer(message, "Товар не найден. Попробуйте уточнить запрос.")
        return

    logger.info(
        "Local catalog fuzzy ok search_source=%s result_count=%s",
        search_source,
        len(filtered[:5]),
    )
    text = f"Товар <code>{query}</code> не найден, но найдены похожие позиции:\n\n"
    for result in filtered[:5]:
        score = int(result["score"])
        text += (
            f"📦 <b>{result['name']}</b>\n"
            f"• Код: {result['kod']}\n"
            f"• Бренд: {result['brend']} | Вид: {result['vid']}\n"
            f"• Цена: {result['price']} ₽ | Наличие: {format_stock_quantity(result['ostatok'])} | "
            f"Склад: {result['sklad']}\n"
            f"📈 Совпадение: {score}%\n\n"
        )

    await _answer(message, text)


async def _send_instruction(message: Message, **kwargs) -> None:
    text = (
        "<b>Инструкция по использованию</b>\n\n"
        "Введите <b>название товара</b> или <b>артикул</b>, например:\n"
        "<code>CM-107</code>\n"
        "<code>ОМ-350</code>\n\n"
        "Можно искать по конкретному складу, указав город перед товаром:\n"
        "<code>Москва CM-107</code>\n"
        "<code>Екатеринбург ОМ-350</code>\n\n"
        "История запросов хранит последние обращения и позволяет быстро повторить поиск."
    )
    await _answer(message, text, **kwargs)
